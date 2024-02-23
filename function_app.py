import azure.functions as func
import logging
from urllib.parse import urlparse
import requests
import openpyxl
from datetime import datetime
import pandas as pd
import json
from openpyxl import load_workbook
import os
import re
from pymongo import MongoClient
import ast
from io import BytesIO
from urllib.parse import urlparse
import traceback

def fetch_json_from_api(url):
    """
    Fetches JSON data from a given API URL.

    Parameters:
    url (str): The URL of the API endpoint.

    Returns:
    dict: The JSON response from the API, parsed into a dictionary.
    """
    response = requests.get(url)
    response = response.json()
    data = response['data']
    if not data:  
        return 0
    else:
        return json.loads(data)

def fetch_excel_from_blob(blob_url):
    """
    Fetches an Excel file from Blob Storage.

    Parameters:
    blob_url (str): The URL of the Excel file stored in Blob Storage.

    Returns:
    openpyxl.workbook.workbook.Workbook: The workbook object containing the Excel data.
    """
    response = requests.get(blob_url)
    excel_data = response.content
    workbook = load_workbook(BytesIO(excel_data), keep_vba=True,  data_only=True)
    return workbook


def make_columns_unique(df):
    cols = pd.Series(df.columns)
    # Replace None values with a placeholder
    cols = cols.fillna('None')
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index.values.tolist()] = [dup + '_' + str(i) if i != 0 else dup for i in range(sum(cols == dup))]
    df.columns = cols
    return df


def create_custom_header(df, row_indices, header_placeholder='Header'):
    indices = list(map(int, row_indices.split(',')))
    custom_header = []

    for column in df.columns:
        values = [df.at[index, column] for index in indices if pd.notna(df.at[index, column])]
        header = ' '.join([str(value) for value in values]) if len(values) > 1 else str(values[0]) if values else None

        # Replace None values with the specified placeholder
        header = header if header is not None else header_placeholder

        custom_header.append(header)

    df.columns = custom_header
    df.drop(indices, inplace=True)
    return df


def column_range(start, end):
    """ Generate a range of Excel column names between start and end. """
    def excel_column_name(n):
        """ Convert a column number to a column name (e.g., 1 -> A, 27 -> AA). """
        name = ''
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            name = chr(65 + remainder) + name
        return name

    start_index = excel_column_number_to_index(start)
    end_index = excel_column_number_to_index(end)
    return [excel_column_name(i) for i in range(start_index, end_index + 1)]

def excel_column_number_to_index(column_name):
    """ Convert a column name to a number (e.g., A -> 1, AA -> 27). """
    sum = 0
    for letter in column_name:
        sum = sum * 26 + (ord(letter) - 64)
    return sum

def parse_column_mappings(mapping_str):
    try:
        mappings = ast.literal_eval(mapping_str)
        for key, value in mappings.items():
            if isinstance(value, list):
                if len(value) == 1:
                    # Single column mapping
                    mappings[key] = value
                elif len(value) == 2:
                    # Range mapping
                    start, end = value.split(':')
                    mappings[key] = column_range(start, end)
            else:
                raise ValueError("Invalid column mapping format. Please provide single column mappings only.")

           
        return mappings
    except (SyntaxError, ValueError):
        # print("Error parsing column mappings.")
        return None


def transform_and_concatenate(df, col_row_index):
    transformed_dfs = []

    for header_col, value_cols in col_row_index.items():
        # Adjust to only include available columns
        available_cols = [col for col in value_cols if col in df.columns]

        # Extract the data for the header column and its corresponding available values
        header = df[header_col].dropna()  # Drop any NaN values in the header column
        values = df[available_cols].loc[df[header_col].notna()]  # Corresponding values, excluding rows where header is NaN

        # Transpose the values and set the header column as the header
        transposed_df = pd.DataFrame(values.values.T, columns=header)
        transformed_dfs.append(transposed_df)

    # Concatenate all transformed dataframes
    concatenated_df = pd.concat(transformed_dfs, axis=1)

    return concatenated_df



def linear_mapping(input_path, sheet_entry, sheet_index, sheet_name, column_mappings, file_id, version_id, project_name, property_name, property_id, post_month, database_name):
    # Load the Excel workbook
    workbook = fetch_excel_from_blob(input_path)

    # Select the specific worksheet by index
    worksheet = workbook.worksheets[sheet_index]

    # Create a list to store the cell info
    cell_info = []

    # Get the start_row from the mapping entry
    start_row = sheet_entry.get('start_row', 1)
    start_row = int(start_row)

    # Iterate through the rows using the specified column mappings
    for col, col_range in column_mappings.items():
        for row_num in range(start_row, worksheet.max_row + 1):
            cell_value = worksheet[f'{col}{row_num}'].value
            if cell_value is not None:
                data_row = {col: cell_value}
                for col_letter in col_range:
                    cell_value = worksheet[f'{col_letter}{row_num}'].value
                    if cell_value is not None:
                        # Convert datetime objects to string
                        if isinstance(cell_value, datetime):
                            cell_value = cell_value.isoformat()
                        data_row[col_letter] = cell_value
                cell_info.append(data_row)

    # Convert the cell_info list into a Pandas DataFrame
    df = pd.DataFrame(cell_info)

    # df.to_csv("new.csv")

    df = transform_and_concatenate(df, column_mappings)

    df = make_columns_unique(df)
    
    # Save the DataFrame to a JSON file
    df_dict = df.to_dict(orient='records')
    output_dict = {
        "FileId" : file_id,
        "VersionId" : version_id,
        "ProjectName": project_name,
        "PropertyName": property_name,
        "Project_id": property_id,
        "Postmonth": post_month,
        "data": df_dict
    }

    # Insert data into MongoDB with project name or ID
    specific_collection_name = sheet_entry.get('collection_name', None)
    insert_into_mongodb(df, database_name, specific_collection_name, file_id, version_id, project_name, property_name, property_id, post_month)
    # print(f"Data inserted into MongoDB collection '{specific_collection_name}'")

    return df, specific_collection_name



def tabular_mapping(input_path, sheet_entry, sheet_index, col_row_index, header_indices, keywords, file_id, version_id, project_name, property_name, property_id, post_month, database_name,required_columns, atleast_header):

    def adjust_column_range(original_range):
        return [original_range[0], original_range[1]]

    def delete_rows_with_keywords(df, keywords, column_range):
        rows_to_delete = set()
        for index, row in df.iterrows():
            for col in range(column_range[0], column_range[1] + 1):
                cell_content = str(row[col]).lower()  # Convert cell content to lower case
                if any(keyword.lower() in cell_content for keyword in keywords):  # Convert keyword to lower case
                    rows_to_delete.add(index)
        return df.drop(rows_to_delete)

   

    def generate_column_names_from_list(header_list):
        custom_header = []
        for column_values in zip(*header_list): # Transpose the header list
            column_values = [value for value in column_values if pd.notna(value)]
            # Join remaining values with a space
            header = ' '.join(str(value) for value in column_values) if column_values else None
            custom_header.append(header)
        return custom_header


    def actual_generate_column_names_from_list(header_list):
        custom_header = []
        for column_values in zip(*header_list):  # Transpose the header list
            unique_values = set(filter(pd.notna, column_values))
            header = ' '.join(str(value) for value in unique_values)
            custom_header.append(header)
        return custom_header

    

    def normalize_unicode_characters(text):
        # Normalize full-width characters and other special Unicode characters to their closest ASCII representation
        # This includes converting full-width spaces (\u3000) and handling newlines
        normalized_text = unicodedata.normalize('NFKC', text)
        # Additional custom replacements can be added here if necessary
        return normalized_text


    
    def generate_column_names(df_slice, row_indices):
        custom_header = []
        actual_row_indices = df_slice.index.tolist()  # Get the actual row indices of the slice
        for column in df_slice.columns:
            values = []
            for i in row_indices:
                if pd.notna(df_slice.at[actual_row_indices[i], column]):
                    value = str(df_slice.at[actual_row_indices[i], column])
                    # Normalize value for all types of Unicode characters
                    normalized_value = unicodedata.normalize('NFKC', value)
                    # Replace newline characters with spaces
                    clean_value = normalized_value.replace('\n', ' ')
                    values.append(clean_value)
            header = ' '.join(values) if values else None
            custom_header.append(header)
        return custom_header


    def remove_empty_rows(df):
            """
            Remove rows where all values across all columns are empty.

            Parameters:
                df (pd.DataFrame): Input DataFrame.

            Returns:
                pd.DataFrame: DataFrame with empty rows removed.
            """
            # Drop rows where all values across all columns are empty
            df_cleaned = df.dropna(how='all')
            
            # Reset index after dropping rows
            df_cleaned.reset_index(drop=True, inplace=True)
            
            return df_cleaned
    
    def extract_and_combine_tables_multirow_header(df, header_row_indices, keywords,required_columns, atleast_header):


        def remove_rows_with_blank_values(df, column_names):
            """
            Remove rows with blank values in the specified columns.

            Parameters:
                df (pd.DataFrame): Input DataFrame.
                column_names (list): List of column names to check for blank values.

            Returns:
                pd.DataFrame: DataFrame with rows removed where any of the specified columns have blank values.
            """
            # Check if df is a DataFrame
            if not isinstance(df, pd.DataFrame):
                # print("Input 'df' is not a DataFrame.")
                # print(df)
                return df
            
            # Ensure column_names is a list
            if not isinstance(column_names, list):
                column_names = [column_names]
            
            # Initialize a boolean mask to select rows with blank values
            mask = df[column_names[0]].isna() | (df[column_names[0]] == '')
            
            # Iterate over remaining column names and update the mask
            for column_name in column_names[1:]:
                mask |= df[column_name].isna() | (df[column_name] == '')
            
            # Find indices of rows with blank values in any of the specified columns
            blank_indices = df[mask].index
            
            # Drop rows with blank values in any of the specified columns
            df_cleaned = df.drop(blank_indices)
            
            # Reset index after dropping rows
            df_cleaned.reset_index(drop=True, inplace=True)
            
            return df_cleaned

        def delete_rows_with_keywords(df, keywords):
            rows_to_delete = set()
            for index, row in df.iterrows():
                for col_name, cell_content in row.items():
                    cell_content = str(cell_content).lower()  # Convert cell content to lower case
                    if any(str(keyword).lower() in cell_content for keyword in keywords):  # Convert keyword to lower case
                        rows_to_delete.add(index)
                        break  # Stop searching for keywords in this row once one is found
            return df.drop(rows_to_delete)
        

        def format_value(value):
                # If the value is a datetime, format it as 'YYYY/MM'
            if isinstance(value, datetime):
                return value.strftime('%Y/%m')
            # If the value is a float but effectively an integer, format it as an integer
            elif isinstance(value, float) and value.is_integer():
                return str(int(value))
            # For floats that are not effectively integers, keep them as float and format to string with desired precision
            elif isinstance(value, float):
                return f"{value:.2f}".rstrip('0').rstrip('.')  # Adjust precision as needed
            # Otherwise, convert the value to string and remove newlines (periods are not removed globally to preserve float formatting)
            else:
                return str(value).replace('\n', '').replace('.', '')

        def generate_column_names_from_list(header_list):
            custom_header = []
            for column_values in zip(*header_list):  # Transpose the header list
                if all(value is None for value in column_values):
                    continue  # Skip this column as all values are None
                # Filter out None values and prepare them
                column_values = [format_value(value) for value in column_values if pd.notna(value)]
                # Join remaining values with a space
                header = ' '.join(column_values) if column_values else None
                custom_header.append(header)
            return custom_header

        def generate_column_names(df_slice, row_indices):
            custom_header = []
            actual_row_indices = df_slice.index.tolist()
            for column in df_slice.columns:
                # Extract and prepare values for each column
                values = [format_value(df_slice.at[actual_row_indices[i], column])
                        for i in row_indices if pd.notna(df_slice.at[actual_row_indices[i], column])]
                header = ' '.join(values) if values else None
                custom_header.append(header)
            return custom_header
        
        
        
        actual_header=generate_column_names_from_list(header_row_indices)
        # print(actual_header)

        index = 0
        matched_header_start_index=None
        comparison_headers=None
        while index < len(df):
            if index + len(header_row_indices) <= len(df):
                comparison_headers = generate_column_names(df.iloc[index:index + len(header_row_indices)], list(range(len(header_row_indices))))
                match_count = sum(ch in comparison_headers for ch in actual_header)

                if match_count>=atleast_header:
                    matched_header_start_index=index
                    break
            
            index+=1

        if matched_header_start_index is not None:
            df.columns = comparison_headers
            df = df.iloc[matched_header_start_index + len(header_indices):].reset_index(drop=True) 
            df=remove_rows_with_blank_values(df,required_columns)
            df=delete_rows_with_keywords(df,keywords)
            # print(df)
            # df.to_csv('df1.csv')

        return df


    def process_sheet(sheet, col_row_index):
        data = []
        for row in sheet.iter_rows(min_row=col_row_index):
            row_data = []
            for cell in row:
                if cell.coordinate in sheet.merged_cells:
                    for range_ in sheet.merged_cells.ranges:
                        if cell.coordinate in range_:
                            value = sheet[range_.start_cell.coordinate].value
                            break
                else:
                    value = cell.value
                row_data.append(value)
            data.append(row_data)
        return data

    # try:
    workbook = fetch_excel_from_blob(input_path)
    df = pd.read_excel(input_path, sheet_name=sheet_index, header=None)
    # print(df.head(10))
    # df.to_csv('df.csv')
    df = extract_and_combine_tables_multirow_header(df, header_indices, keywords,required_columns, atleast_header)
    df=remove_empty_rows(df)
    df.columns = df.columns.astype(str).str.lower()
    df = make_columns_unique(df)

        # Save the DataFrame to a JSON file
    output_dict = {
        "FileId" : file_id,
        "VersionId" : version_id,
        "ProjectName": project_name,
        "PropertyName": property_name,
        "Project_id": property_id,
        "Postmonth": post_month,
        "data": df.to_dict(orient='records'),
        "date_format": "iso",
        "default_handler": str
        }
    # Insert data into MongoDB with project name or ID
    specific_collection_name = sheet_entry.get('collection_name', None)
    insert_into_mongodb(df, database_name, specific_collection_name, file_id, version_id, project_name, property_name, property_id, post_month)


    return df, specific_collection_name
    # except Exception as e:
    #     # print("An error occurred:", e)
    #     return None, None





def mt1_mapping(input_path, sheet_entry, sheet_index, col_row_index, rows_col_to_exclude, header_indices, columns_added, keywords, column_name, del_type, atleast_header, file_id, version_id, project_name, property_name, property_id, post_month, database_name):

    def adjust_column_range(original_range, columns_added):
        return [original_range[0] + columns_added, original_range[1] + columns_added]

    def delete_rows_with_keywords(df, keywords, column_range):
        rows_to_delete = set()
        for index, row in df.iterrows():
            for col in range(column_range[0], column_range[1] + 1):
                cell_content = str(row[col]).lower()  # Convert cell content to lower case
                if any(keyword.lower() in cell_content for keyword in keywords):  # Convert keyword to lower case
                    rows_to_delete.add(index)
        return df.drop(rows_to_delete)

    def generate_column_names(df_slice, row_indices):
        custom_header = []
        actual_row_indices = df_slice.index.tolist()  # Get the actual row indices of the slice
        for column in df_slice.columns:
            values = [df_slice.at[actual_row_indices[i], column] for i in row_indices if pd.notna(df_slice.at[actual_row_indices[i], column])]
            header = ' '.join([str(value) for value in values]) if values else None
            custom_header.append(header)
        return custom_header
    
    def actual_generate_column_names(df_slice, row_indices):
        custom_header = []
        actual_row_indices = df_slice.index.tolist()  # Get the actual row indices of the slice
        for column in df_slice.columns:
            # Use a set to store unique values for each column
            unique_values = set()
            for i in row_indices:
                value = df_slice.at[actual_row_indices[i], column]
                if pd.notna(value):
                    unique_values.add(value)
            # Join the unique values to form the header
            header = ' '.join(str(value) for value in unique_values)
            custom_header.append(header)
        return custom_header

    def extract_and_combine_tables_multirow_header(df, header_row_indices, keywords, column_range, column_name, atleast_header):
        header_row_indices = [i - 1 for i in header_row_indices]
        column_headers = generate_column_names(df.iloc[header_row_indices], list(range(len(header_row_indices))))
        actual_column_headers = actual_generate_column_names(df.iloc[header_row_indices], list(range(len(header_row_indices))))

        combined_data = []
        current_table_data = []
        in_table = False
        index = 0
        current_type_value = None  # This will hold the 'type' value for the current section
        type_column_name = column_name  # Default name for the 'Type' column

        while index < len(df):
            if index + len(header_row_indices) <= len(df):
                comparison_headers = generate_column_names(df.iloc[index:index + len(header_row_indices)], list(range(len(header_row_indices))))
                match_count = sum(ch == comp_ch for ch, comp_ch in zip(column_headers, comparison_headers) if pd.notnull(comp_ch))

                if match_count >= atleast_header:
                    # Retrieve and concatenate all non-empty values from the row above the header
                    type_row = df.iloc[index - 1].dropna()
                    current_type_value = ' '.join(map(str, type_row.values))

                    if current_table_data:
                        for row in current_table_data:
                            row.insert(0, current_type_value)
                        combined_data.extend(current_table_data)
                        current_table_data = []
                    in_table = True
                    index += len(header_row_indices)  # Skip header rows
                    continue

            if in_table and df.iloc[index].isnull().all():
                if current_table_data:
                    combined_data.extend(current_table_data)
                    current_table_data = []
                in_table = False

            elif in_table:
                current_row = df.iloc[index].tolist()
                current_row.insert(0, current_type_value)  # Add the 'type' value to the beginning of the row
                current_table_data.append(current_row)

            index += 1

        if current_table_data:
            # for row in current_table_data:
            #     row.insert(0, current_type_value)
            combined_data.extend(current_table_data)

        # Adjust the column headers to include the new 'Type' column
        actual_column_headers.insert(0, type_column_name)
        # print("actual_column_headers",actual_column_headers)
        combined_table = pd.DataFrame(combined_data)
        # print("combined_data",combined_data)


        combined_table = pd.DataFrame(combined_data, columns=actual_column_headers)

        # Keep only the first header for the entire DataFrame
        # combined_table.columns = combined_table.iloc[0].tolist()
        # combined_table.to_csv('combined.csv')


        combined_table = delete_rows_with_keywords(combined_table, keywords, column_range)
        return combined_table



    def process_sheet(sheet,col_row_index):
        """
        Process a single sheet to handle merged cells, writing the values instead of formulas,
        and skipping the first 7 rows.
        """
        data = []
        for row in sheet.iter_rows(min_row=col_row_index):  # Start from the 8th row, skipping the first 7
            row_data = []
            for cell in row:
                if cell.coordinate in sheet.merged_cells:
                    # For a merged cell, find its range and get the value
                    for range_ in sheet.merged_cells.ranges:
                        if cell.coordinate in range_:
                            value = sheet[range_.start_cell.coordinate].value
                            break
                else:
                    value = cell.value
                row_data.append(value)
            data.append(row_data)
        return data
    
    
    workbook = load_workbook(input_path, data_only=True)

    if 0 <= sheet_index < len(workbook.worksheets):
        sheet = workbook.worksheets[sheet_index]
        data = process_sheet(sheet, col_row_index)
        df = pd.DataFrame(data)
        # df.to_csv('process_sheet.csv')
        rows_col_to_exclude = [i - 1 for i in rows_col_to_exclude]
        adjusted_column_range = adjust_column_range(rows_col_to_exclude, columns_added)
        df = extract_and_combine_tables_multirow_header(df, header_indices,keywords, adjusted_column_range,column_name,atleast_header)
        # df.to_csv("df.csv")
        df=make_columns_unique(df)
        # df.to_csv("df1.csv")
        if del_type == 0:
            df.drop(columns=column_name, inplace=True)

    


        # Convert DataFrame to JSON

        df_dict = df.to_dict
        output_dict = json.dumps({
            "FileId" : file_id,
            "VersionId" : version_id,
            "ProjectName": project_name,
            "PropertyName": property_name,
            "Project_id": property_id,
            "Postmonth": post_month,
            "data": df.to_dict(orient='records', lines=False)
            }, indent=4, default=str)
            


    specific_collection_name = sheet_entry.get('collection_name', None)
    insert_into_mongodb(df, database_name, specific_collection_name, file_id, version_id, project_name, property_name, property_id, post_month)
    # print(f"Data inserted into MongoDB collection '{specific_collection_name}'")

    return df, specific_collection_name



def lmt_mapping(input_path, sheet_entry, sheet_index, col_row_index, file_id, version_id, project_name, property_name, property_id, post_month, database_name):

    def delete_rows_with_keywords(df, keywords):
        rows_to_delete = set()
        for index, row in df.iterrows():
            for col in df.columns:
                cell_content = str(row[col]).lower()  # Convert cell content to lower case
                if any(keyword.lower() in cell_content for keyword in keywords):  # Check if any keyword is in the cell content
                    rows_to_delete.add(index)
                    break  # Break the inner loop if a keyword is found
        return df.drop(rows_to_delete)

    def read_excel_range(input_path, sheet_index, range_str):
        """
        Reads a specific range from an Excel sheet using openpyxl.

        Args:
        input_path (str): Path to the Excel file.
        sheet_index (int): Index of the sheet to read from (0-indexed).
        range_str (str): The column range in the format 'Start:End' (e.g., 'B:L').

        Returns:
        list of list: Data read from the specified range in the Excel sheet.
        """
        from openpyxl import load_workbook

        # Extract start and end columns from the range string
        start_column, end_column = range_str.split(':')

        # Load the workbook
        wb = load_workbook(filename=input_path, read_only=True, data_only=True)

        # Get the sheet by index
        sheet = wb.worksheets[sheet_index]

        # Define the row range
        start_row = 1  # Starting from the first row
        end_row = sheet.max_row  # Read till the last row of the sheet

        # Read data from the range
        data = []
        for row in sheet[f'{start_column}{start_row}':f'{end_column}{end_row}']:
            row_data = [cell.value for cell in row]
            data.append(row_data)

        df = pd.DataFrame(data)

        return df
    
    def preprocess_string(s):
        """
        Preprocesses a string by making it lowercase and removing spaces.

        Args:
        s (str): The string to preprocess.

        Returns:
        str: The preprocessed string.
        """
        return ''.join(s.lower().split())

    def extract_and_combine_tables(df, header, atleast_header):
        """
        Extracts and combines tables from a DataFrame based on a multi-row header.

        Args:
        df (DataFrame): The DataFrame to process.
        header (list): The header to match for identifying the start of a table.
        atleast_header (int): Minimum number of header elements that must match to consider it the start of a table.

        Returns:
        list of list: Combined data from all identified tables.
        """
        # Preprocess header elements
        preprocessed_header = [preprocess_string(h) for h in header]
        combined_data = []
        column_headers = None
        current_table_data = []
        in_table = False
        index = 0

        while index < len(df):
            row = df.iloc[index]

            # Preprocess row elements and check if the row matches the header
            preprocessed_row = [preprocess_string(str(cell)) for cell in row]
            match_count = sum(cell in preprocessed_header for cell in preprocessed_row)

            if match_count >= atleast_header:
                current_row = row.tolist()
                current_table_data.append(current_row)
                # Start of a new table
                # if current_table_data:
                #     combined_data.extend(current_table_data)
                #     current_table_data = []
                in_table = True
                index += 1  # Skip the header row
                continue

            # Check for the end of the table (empty row)
            if in_table and row.isnull().all():
                if current_table_data:
                    combined_data.extend(current_table_data)
                    current_table_data = []
                in_table = False
                break

            # If in a table, add the row to the current table data
            elif in_table:
                current_row = row.tolist()
                current_table_data.append(current_row)

            index += 1

        # Add the last table data if any
        if current_table_data:
            combined_data.extend(current_table_data)

        combined_df=pd.DataFrame(combined_data)

        return combined_df

    for col_row in col_row_index:
        range=col_row['range']
        tables=col_row['tables']
        for table in tables:
            header=table['header']
            keywords=table['keywords']
            collection_name=table['collection_name']
            atleast_header=len(header)
            data = read_excel_range(input_path, sheet_index, range)
            df=extract_and_combine_tables(data,header,atleast_header)
            # Drop columns where all values are NaN
            df = df.dropna(axis=1, how='all')
            # Drop rows where all values are NaN
            df = df.dropna(axis=0, how='all')
            # Set the first row as the header
            df.columns = df.iloc[0]
            # Drop the first row
            df = df.drop(df.index[0])
            df=delete_rows_with_keywords(df,keywords)
        

            df_dict = df.to_dict
            output_dict = json.dumps({
                "FileId" : file_id,
                "VersionId" : version_id,
                "ProjectName": project_name,
                "PropertyName": property_name,
                "Project_id": property_id,
                "Postmonth": post_month,
                "data": df.to_dict(orient='records', lines=False)
                }, indent=4, default=str)


            insert_into_mongodb(df, database_name, collection_name, file_id, version_id, project_name, property_name, property_id, post_month)
            # print(f"Data inserted into MongoDB collection '{collection_name}'") 

    

# Function to insert data into MongoDB
def insert_into_mongodb(df, database_name, collection_name, file_id, version_id, project_name, property_name, property_id, post_month):
    # client = MongoClient('mongodb+srv://mahendrapatel:llDpDNfm3BBP0H4q@cluster0.1yrytz8.mongodb.net/?retryWrites=true&w=majority')  # Connect to your local MongoDB server
    client = MongoClient('mongodb://localhost:27017')   
    db = client[database_name]  # Use the provided database name
    collection = db[collection_name]

    # Ensure unique column names
    df.columns = pd.Index([str(col) for col in df.columns])

    # Convert DataFrame to dictionary and insert into MongoDB
    # df_json = df.astype(str).to_json(orient='records')

    # Convert JSON string back to dictionary for insertion
    # records = json.loads(df_json)
    records = df.to_dict(orient='records')

    if not records:
        # print("No records to insert into MongoDB.")
        return

    for record in records:
        record.update({
            "FileId" : file_id,
            "VersionId" : version_id,
            "ProjectName": project_name,
            "PropertyName": property_name,
            "Project_id": property_id,
            "Postmonth": post_month
        })

    collection.insert_many(records)
    


def process_mapping(mapping_data):
    file_id = None
   
    for mapping_entry in mapping_data:
        file_id = (mapping_entry.get('FileId'))
        version_id = (mapping_entry.get("VersionId"))
        # print(mapping_entry)
        
        input_path = mapping_entry.get('File_path', '')
        project_name = mapping_entry.get('ProjectName', '')
        property_name = mapping_entry.get('PropertyName', '')
        property_id = mapping_entry.get('PropertyId', '')
        post_month = mapping_entry.get('PostMonth', '')
        database_name = mapping_entry.get('DatabaseName', 'PAGStaging') 

        workbook = fetch_excel_from_blob(input_path)
        sheet_names = workbook.sheetnames

        for sheet_entry in mapping_entry['Sheets']:
            sheet_index = int(sheet_entry['sheet_index'])
            sheet_name = sheet_names[sheet_index]
            sheet_type = sheet_entry['sheet_type']

            skip_sheet = sheet_entry.get('skip', 0)  # Default to 0 (not skipped) if not provided
            skip_sheet = int(skip_sheet)  # Convert to integer

            if skip_sheet:
            
                continue

            if sheet_type == 'L':
                
                col_row_index_str = sheet_entry.get('col_row_index', '')
                column_mappings = parse_column_mappings(col_row_index_str)

                if column_mappings is not None:
                    linear_mapping_result, specific_collection_name = linear_mapping(
        input_path, sheet_entry, sheet_index, sheet_name, column_mappings, file_id, version_id,
        project_name, property_name, property_id, post_month, database_name)
                    

            elif sheet_type == 'T':
                header_indices = sheet_entry['headers']
                col_row_index = int(sheet_entry['col_row_index'])
                required_columns = sheet_entry['required']
                atleast_header = int(sheet_entry['atleast_header'])
                # rows_col_to_exclude = sheet_entry['rows_col_to_exclude']
                keywords = sheet_entry['keywords']
                

                result_df, specific_collection_name = tabular_mapping(input_path, sheet_entry, sheet_index, col_row_index, header_indices, keywords, file_id, version_id, project_name, property_name, property_id, post_month, database_name, required_columns, atleast_header) 
                

            elif sheet_type == 'MT':
                col_row_index = sheet_entry['col_row_index']
                rows_col_to_exclude = sheet_entry['rows_col_to_exclude']
                output_path_json = f'mt_sheet_{sheet_index}.json'

                result_df, specific_collection_name = mt_mapping(input_path, sheet_entry, sheet_index, sheet_name, col_row_index, rows_col_to_exclude, file_id, version_id, project_name, property_name, property_id, post_month, database_name)
                

            elif sheet_type == 'MTC':
                # sheet_index = int(sheet_entry['sheet_index'])
                col_row_index = int(sheet_entry['col_row_index'])
                rows_col_to_exclude = sheet_entry['rows_col_to_exclude']
                header_indices = sheet_entry['header']
                # columns_added = sheet_entry['col_add']
                columns_added=1
                keywords = sheet_entry['keywords']
                column_name = sheet_entry['col_name']
                output_path_json = f'mt_sheet_{sheet_index}.json'
                atleast_header=sheet_entry['atleast_header']
                del_type = sheet_entry.get('del_type', 0)
                result_df, specific_collection_name = mt1_mapping(input_path, sheet_entry, sheet_index, col_row_index, rows_col_to_exclude, header_indices, columns_added, keywords, column_name, del_type, atleast_header, file_id, version_id, project_name, property_name, property_id, post_month, database_name)
                

            elif sheet_type == 'LMT':
                col_row_index = sheet_entry['col_row_index']
                lmt_mapping(input_path,sheet_entry, sheet_index,col_row_index, file_id, version_id, project_name, property_name, property_id, post_month, database_name)
        # break
    return file_id


app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)

# @app.route(route="http_trig_mapping")
@app.route(route="http_trig_mapping_dap_demo")
def http_trig_mapping_dap_demo(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')
    
    api_url = "https://dap-demo-api.4seeanalytics.com/api/fileworkflow/submission/datafor/validationjob"
    if not api_url:
        return func.HttpResponse("API URL not provided", status_code=400)

    # Fetch JSON data from API URL
    api_data = fetch_json_from_api(api_url)
    # file_id = None
    # for mapping_entry in mapping_data:
    #     file_id = mapping_entry.get('FileId')

    if  api_data != 0 :

    # # Extract input file path (Blob URL) and mapping JSON from API data
        # blob_url = api_data.get('File_path')

        if api_data:
            # Process mapping 
            file_id = process_mapping(api_data)

            # Perform further operations with mapped data
            json_response = json.dumps({"fileId": file_id, "process_status": 1, "status_code": 200}, indent=4)
            return func.HttpResponse(json_response, status_code=200)
        else:
            error_json={
                "error":"Blob URL or Mapping JSON not found in API data",
                "process_status": 0,
                "status_code":400
            }
            return func.HttpResponse(json.dumps(error_json,indent=4),status_code=400)
            # return func.HttpResponse("Blob URL or Mapping JSON not found in API data", status_code=400)
    else:
        error_json={
            "error":"Nothing to process",
            "process_status": 0,
            "status_code":200
        }
        return func.HttpResponse(json.dumps(error_json,indent=4),status_code=200)




